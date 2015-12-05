import argparse

import openpyxl

from . import las


class ExcelConverter(object):
    '''
    Arguments:
        las: LASFile object

    '''

    def __init__(self, las):
        self.las = las

    def set_las(self, las):
        self.las = las

    def write(self, xlsxfn):
        assert xlsxfn.lower().endswith('.xlsx')
        wb = openpyxl.Workbook()
        header = wb['Sheet']
        # header = wb.create_sheet()
        header.title = 'Header'
        curves = wb.create_sheet()
        curves.title = 'Curves'

        def write_cell(sh, i, j, value):
            c = sh.cell(row=i + 1, column=j + 1)
            c.value = value

        write_cell(header, 0, 0, "Section")
        write_cell(header, 0, 1, "Mnemonic")
        write_cell(header, 0, 2, "Unit")
        write_cell(header, 0, 3, "Value")
        write_cell(header, 0, 4, "Description")

        sections = [
            ('~Version', self.las.version),
            ('~Well', self.las.well),
            ('~Parameter', self.las.params)
            ]

        n = 1
        for sect_name, sect in sections:
            for i, item in enumerate(sect.values()):
                write_cell(header, n, 0, sect_name)
                write_cell(header, n, 1, item.mnemonic)
                write_cell(header, n, 2, item.unit)
                write_cell(header, n, 3, item.value)
                write_cell(header, n, 4, item.descr)
                n += 1

        for i, curve in enumerate(self.las.curves):
            write_cell(curves, 0, i, curve.mnemonic)
            for j, value in enumerate(curve.data):
                write_cell(curves, j + 1, i, value)

        wb.save(xlsxfn)


def main():
    args = get_parser().parse_args(sys.argv[1:])
    lasfn = args.LAS_filename
    xlsxfn = args.XLSX_filename

    l = las.LASFile(lasfn)
    converter = ExcelConverter(l)
    converter.write_excel(xlsxfn)


def get_parser():
    parser = argparse.ArgumentParser('Convert LAS file to XLSX')
    parser.add_argument('LAS_filename')
    parser.add_argument('XLSX_filename')
    return parser


if __name__ == '__main__':
    main()
